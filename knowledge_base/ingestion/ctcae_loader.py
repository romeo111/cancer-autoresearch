"""CTCAE v5.0 loader.

Reads NCI's official CTCAE v5.0 Excel workbook
(https://evs.nci.nih.gov/ftp1/CTCAE/CTCAE_5.0/CTCAE_v5.0_2017-11-27.xlsx)
or a CSV with similar column shape.

Output: knowledge_base/hosted/ctcae/v5.0/grading.yaml
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Iterable

import yaml


def _normalize_header(s: str) -> str:
    """'Grade 1 \\xa0\\xa0' → 'grade_1' so we have stable keys."""
    s = (s or "").replace("\xa0", " ").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s


def _row_to_entry(row: dict) -> dict | None:
    term = (row.get("ctcae_term") or "").strip()
    if not term:
        return None
    soc = (row.get("meddra_soc") or "").strip() or None
    code_slug = re.sub(r"[^a-z0-9]+", "_", term.lower()).strip("_")
    meddra_code = (row.get("meddra_code") or "").strip() or None
    return {
        "code": f"CTCAE.{code_slug}",
        "meddra_code": meddra_code,
        "term": term,
        "soc_category": soc,
        "grades": {
            "1": (row.get("grade_1") or "").strip() or None,
            "2": (row.get("grade_2") or "").strip() or None,
            "3": (row.get("grade_3") or "").strip() or None,
            "4": (row.get("grade_4") or "").strip() or None,
            "5": (row.get("grade_5") or "").strip() or None,
        },
        "definition": (row.get("definition") or "").strip() or None,
        "navigational_note": (row.get("navigational_note") or "").strip() or None,
    }


def parse_ctcae_csv(rows: Iterable[dict]) -> list[dict]:
    aes: list[dict] = []
    for raw in rows:
        norm = {_normalize_header(k): v for k, v in raw.items()}
        e = _row_to_entry(norm)
        if e is not None:
            aes.append(e)
    return aes


def parse_ctcae_xlsx(xlsx_path: Path, sheet_name: str = "CTCAE v5.0 Clean Copy") -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        # Fall back to first sheet
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    headers_raw = next(rows_iter)
    headers = [_normalize_header(str(h or "")) for h in headers_raw]

    aes: list[dict] = []
    for row in rows_iter:
        d = {headers[i]: ("" if v is None else str(v)) for i, v in enumerate(row) if i < len(headers)}
        e = _row_to_entry(d)
        if e is not None:
            aes.append(e)
    return aes


def load_ctcae(input_path: Path, out_dir: Path) -> Path:
    if input_path.suffix.lower() in {".xlsx", ".xlsm"}:
        aes = parse_ctcae_xlsx(input_path)
    else:
        with input_path.open(encoding="utf-8", newline="") as f:
            aes = parse_ctcae_csv(csv.DictReader(f))

    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "grading.yaml"
    payload = {
        "source_id": "SRC-CTCAE-V5",
        "version": "v5.0",
        "source_url": "https://evs.nci.nih.gov/ftp1/CTCAE/CTCAE_5.0/CTCAE_v5.0_2017-11-27.xlsx",
        "license": "Public domain (NCI)",
        "adverse_events": aes,
    }
    out_path.write_text(
        yaml.safe_dump(payload, sort_keys=False, allow_unicode=True),
        encoding="utf-8",
    )
    return out_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Load CTCAE v5.0 (xlsx or csv) into YAML.")
    parser.add_argument("input", type=Path, help="CTCAE v5.0 .xlsx or .csv")
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=Path("knowledge_base/hosted/ctcae/v5.0"),
    )
    args = parser.parse_args()

    if not args.input.is_file():
        print(f"ERROR: input not found: {args.input}", file=sys.stderr)
        return 2

    out = load_ctcae(args.input, args.out_dir)
    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
